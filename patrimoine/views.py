from django.core.paginator import Paginator
from django.shortcuts import render

# Create your views here.

# importation de toutes les bibliothèques nécessaire
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect, HttpResponse
from django.db.models import Sum, F, Value
from django.urls import reverse
from openpyxl.styles import Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Spacer, PageBreak
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import HttpResponse
import io
from django.contrib import messages
from openpyxl import Workbook
from .models import PVRecensement, Inventaire, EtatApreciatif, OrdreEntree, OrdreSortie, AttestationPriseEnCharge, \
    Region, Service, PvEvaluation


# C'est ici qu'on va crée tout les fonctions


def ajout_region(request):
    if request.method == "POST":
        Nom = request.POST.get("nom")
        Region.objects.create(Nom=Nom)
        return redirect('afficher_regions')
    return render(request, "patrimoine/ajout_region.html")


def ajout_services(request, region_id):
    region = get_object_or_404(Region, id=region_id)
    if request.method == "POST":
        Nom = request.POST.get("nom")
        Service.objects.create(Region=region, Nom=Nom)  # Utilisation de l'objet Region
        return redirect(reverse('afficher_services', args=[region_id]))
    return render(request, 'patrimoine/ajout_services.html', {'region': region})


def afficher_regions(request):
    regions = Region.objects.all()
    return render(request, 'patrimoine/regions.html', {'regions': regions})


def afficher_services(request, region_id):
    region = get_object_or_404(Region, id=region_id)
    services = Service.objects.filter(Region=region)  # Utilisation du champ 'Region'
    return render(request, 'patrimoine/services.html', {'region': region, 'services': services})


# Page d'accueil, veux-dire, dès qu'on demarre l'application, c'est cette vue qu'i s'affichera en premier, notament c'est un Login
def home(request):
    if request.method == "POST":
        username = request.POST.get("nom")
        password = request.POST.get("password")
        user = authenticate(username=username, password=password)

        if user:
            login(request, user)
            return redirect('afficher_regions')
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect")
    return render(request, "patrimoine/home.html")


# fonction pour la deconnexion, on utilise @login_required pour interdire à tout personne d'y acceder sans se connecter
@login_required
def Deconnexion(request):
    logout(request)
    return redirect('home')


# fonction d'ajout d'un procès verbal de recensement
@login_required
def ajout_pv_recensement(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Budget = request.POST.get('Budget')
        Imputation_administrative = request.POST.get('Imputation_administrative')
        Designation_chapitre = request.POST.get('Designation_chapitre')
        Libelle_article = request.POST.get('Libelle_article')
        Designation_magasin_ou_materiels_service = request.POST.get('Designation_magasin_ou_materiels_service')
        Nom_et_qualite_recenseur = request.POST.get('Nom_et_qualite_recenseur')
        Nom_et_qualite_depositaire_comptable = request.POST.get('Nom_et_qualite_depositaire_comptable')
        Annee_exercice = request.POST.get("exercice")
        Nomenclature = request.POST.get("Nomenclature")
        Designation_materiels = request.POST.get("Designation_materiels")
        Especes_unites = request.POST.get("especes")
        Prix_unites = request.POST.get("prix")
        Quantites_d_apres_ecriture = request.POST.get("quantite_ecriture")
        Quantites_par_recensement = request.POST.get("quantite_recensement")
        Quantites_execedent_par_article = request.POST.get("excedent_article")
        Quantites_deficient_par_article = request.POST.get("deficient_article")
        valeurs_excedents_par_article = request.POST.get("valeurs_excedents_par_article")
        valeurs_excedents_par_nomenclature = request.POST.get("valeurs_excedents_par_nomenclature")
        valeurs_deficits_par_article = request.POST.get("valeurs_deficits_par_article")
        valeurs_des_deficits_par_nomenclature = request.POST.get("valeurs_des_deficits_par_nomenclature")
        valeurs_des_existants = request.POST.get("valeurs_des_existants")
        Observations = request.POST.get("Observations")

        PVRecensement.objects.create(
            Service=service,
            Budget=Budget,
            Imputation_administrative=Imputation_administrative,
            Designation_chapitre=Designation_chapitre,
            Libelle_article=Libelle_article,
            Designation_magasin_ou_materiels_service=Designation_magasin_ou_materiels_service,
            Nom_et_qualite_recenseur=Nom_et_qualite_recenseur,
            Nom_et_qualite_depositaire_comptable=Nom_et_qualite_depositaire_comptable,
            Annee_exercice=Annee_exercice,
            Nomenclature=Nomenclature,
            Designation_materiels=Designation_materiels,
            Especes_unites=Especes_unites,
            Prix_unites=Prix_unites,
            Quantites_d_apres_ecriture=Quantites_d_apres_ecriture,
            Quantites_par_recensement=Quantites_par_recensement,
            Quantites_execedent_par_article=Quantites_execedent_par_article,
            Quantites_deficient_par_article=Quantites_deficient_par_article,
            valeurs_excedents_par_article=valeurs_excedents_par_article,
            valeurs_excedents_par_nomenclature=valeurs_excedents_par_nomenclature,
            valeurs_deficits_par_article=valeurs_deficits_par_article,
            valeurs_des_deficits_par_nomenclature=valeurs_des_deficits_par_nomenclature,
            valeurs_des_existants=valeurs_des_existants,
            Observations=Observations
        )

        return redirect(reverse('pv_recensement', args=[service_id]))
    return render(request, "patrimoine/ajout_pv_recensement.html", {"service": service})


# fonction d'ajout d'un etat appreciatif
@login_required
def ajout_etat_apreciatif(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Budget = request.POST.get('Budget')
        Materiel_en_approvisionnement_ou_en_service = request.POST.get('Materiel_en_approvisionnement_ou_en_service')
        Designation_du_chapitre = request.POST.get('Designation_du_chapitre')
        libelle_article = request.POST.get('libelle_article')
        Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service = request.POST.get(
            'Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service')
        Nombre_en_toutes_lettres = request.POST.get('Nombre_en_toutes_lettres')
        Nom_et_qualite_du_depositaire_comptable = request.POST.get('Nom_et_qualite_du_depositaire_comptable')
        Annee_exercice = request.POST.get("Annee_exercice")
        Nomenclature = request.POST.get("Nomenclature")
        Numero_du_piece_justificative = request.POST.get("Numero_du_piece_justificative")
        Date_du_piece_justificative = request.POST.get("Date_du_piece_justificative")
        Designations_sommaire_des_operations = request.POST.get("Designations_sommaire_des_operations")
        Charge = request.POST.get("Charge")
        Decharge = request.POST.get("Decharge")

        EtatApreciatif.objects.create(Service=service,
                                      Budget=Budget,
                                      Materiel_en_approvisionnement_ou_en_service=Materiel_en_approvisionnement_ou_en_service,
                                      Designation_du_chapitre=Designation_du_chapitre,
                                      libelle_article=libelle_article,
                                      Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service=Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service,
                                      Nombre_en_toutes_lettres=Nombre_en_toutes_lettres,
                                      Nom_et_qualite_du_depositaire_comptable=Nom_et_qualite_du_depositaire_comptable,
                                      Annee_exercice=Annee_exercice,
                                      Nomenclature=Nomenclature,
                                      Numero_du_piece_justificative=Numero_du_piece_justificative,
                                      Date_du_piece_justificative=Date_du_piece_justificative,
                                      Designations_sommaire_des_operations=Designations_sommaire_des_operations,
                                      Charge=Charge,
                                      Decharge=Decharge)
        return redirect(reverse('etat_appreciatif', args=[service_id]))
    return render(request, "patrimoine/ajout_etat_apreciatif.html", {"service": service})


# fonction d'ajout d'un inventaire
@login_required
def ajout_inventaire(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Budget = request.POST.get('Budget')
        imputation_administrative = request.POST.get('imputation_administrative')
        Designation_du_chapitre = request.POST.get('Designation_du_chapitre')
        libelle_article = request.POST.get('libelle_article')
        Nom_et_qualite_du_recenseur = request.POST.get('Nom_et_qualite_du_recenseur')
        Nom_et_qualite_du_depositaire_comptable = request.POST.get('Nom_et_qualite_du_depositaire_comptable')
        Annee_exercice = request.POST.get("exercice")
        Nomenclature = request.POST.get("Nomenclature")
        Numero_folio_grand_livre = request.POST.get("Numero_folio_grand_livre")
        Designation_materiels = request.POST.get("Designation_materiels")
        Especes_des_unites = request.POST.get("Especes_des_unites")
        Prix_de_l_unite = request.POST.get("Prix_de_l_unite")
        Quantite_existant_1er_janvier = request.POST.get("Quantite_existant_1er_janvier")
        Quantite_entree_pendant_l_annee = request.POST.get("Quantite_entree_pendant_l_annee")
        Quantite_sortie_pendant_l_annee = request.POST.get("Quantite_sortie_pendant_l_annee")
        Quantite_reste_31_decembre = request.POST.get("Quantite_reste_31_decembre")
        Decompte = request.POST.get("Decompte")
        Observation = request.POST.get("Observation")

        Inventaire.objects.create(Service=service,
                                  Budget=Budget,
                                  imputation_administrative=imputation_administrative,
                                  Designation_du_chapitre=Designation_du_chapitre,
                                  libelle_article=libelle_article,
                                  Nom_et_qualite_du_recenseur=Nom_et_qualite_du_recenseur,
                                  Nom_et_qualite_du_depositaire_comptable=Nom_et_qualite_du_depositaire_comptable,
                                  Annee_exercice=Annee_exercice,
                                  Nomenclature=Nomenclature,
                                  Numero_folio_grand_livre=Numero_folio_grand_livre,
                                  Designation_materiels=Designation_materiels,
                                  Especes_des_unites=Especes_des_unites,
                                  Prix_de_l_unite=Prix_de_l_unite,
                                  Quantite_existant_1er_janvier=Quantite_existant_1er_janvier,
                                  Quantite_entree_pendant_l_annee=Quantite_entree_pendant_l_annee,
                                  Quantite_sortie_pendant_l_annee=Quantite_sortie_pendant_l_annee,
                                  Quantite_reste_31_decembre=Quantite_reste_31_decembre,
                                  Decompte=Decompte,
                                  Observation=Observation)
        return redirect(reverse('inventaire', args=[service_id]))
    return render(request, "patrimoine/ajout_inventaire.html", {"service": service})


# fonction d'ajout d'un ordre d'entrée
@login_required
def ajout_ordre_entree(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Budget = request.POST.get('Budget')
        imputation_administrative = request.POST.get('imputation_administrative')
        Designation_du_chapitre = request.POST.get('Designation_du_chapitre')
        libelle_article = request.POST.get('libelle_article')
        Subdivision_du_chapitre = request.POST.get('Subdivision_du_chapitre')
        Numero_du_journal = request.POST.get('Numero_du_journal')
        Nom_et_qualite_du_depositaire_comptable = request.POST.get('Nom_et_qualite_du_depositaire_comptable')
        Annee_exercice = request.POST.get("Annee_exercice")
        Numero_folio_du_grandlivre = request.POST.get("Numero_folio_du_grandlivre")
        Nomenclature = request.POST.get("Nomenclature")
        Designation_des_matieres_et_objets = request.POST.get("Designation_des_matieres_et_objets")
        Especes_des_unites = request.POST.get("Especes_des_unites")
        Quantites = request.POST.get("Quantites")
        Prix_unite = request.POST.get("Prix_unite")
        Valeurs_partielles = request.POST.get("Valeurs_partielles")
        Valeurs_par_numero_nomenclature = request.POST.get("Valeurs_par_numero_nomenclature")
        Numero_piece_justificative_sortie_correspondante = request.POST.get(
            "Numero_piece_justificative_sortie_correspondante")

        OrdreEntree.objects.create(Service=service,
                                   Budget=Budget,
                                   imputation_administrative=imputation_administrative,
                                   Designation_du_chapitre=Designation_du_chapitre,
                                   libelle_article=libelle_article,
                                   Subdivision_du_chapitre=Subdivision_du_chapitre,
                                   Numero_du_journal=Numero_du_journal,
                                   Nom_et_qualite_du_depositaire_comptable=Nom_et_qualite_du_depositaire_comptable,
                                   Annee_exercice=Annee_exercice,
                                   Numero_folio_du_grandlivre=Numero_folio_du_grandlivre,
                                   Nomenclature=Nomenclature,
                                   Designation_des_matieres_et_objets=Designation_des_matieres_et_objets,
                                   Especes_des_unites=Especes_des_unites,
                                   Quantites=Quantites,
                                   Prix_unite=Prix_unite,
                                   Valeurs_partielles=Valeurs_partielles,
                                   Valeurs_par_numero_nomenclature=Valeurs_par_numero_nomenclature,
                                   Numero_piece_justificative_sortie_correspondante=Numero_piece_justificative_sortie_correspondante)

        return redirect(reverse('ordre_entree', args=[service_id]))
    return render(request, "patrimoine/ajout_ordre_entree.html", {"service": service})


# fonction d'ajout d'un attestation de prise en charge
@login_required
def ajout_attestation_prise_en_charge(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Annee_exercice = request.POST.get("Annee_exercice")
        Nomenclature = request.POST.get("Nomenclature")
        Designation_des_matieres_et_objets = request.POST.get("Designation_des_matieres_et_objets")
        Especes_des_unites = request.POST.get("Especes_des_unites")
        Quantite = request.POST.get("Quantite")
        Prix_unite = request.POST.get("Prix_unite")
        Montant = request.POST.get("Montant")
        Observations = request.POST.get("Observations")

        AttestationPriseEnCharge.objects.create(Designation_des_matieres_et_objets=Designation_des_matieres_et_objets,
                                                Nomenclature=Nomenclature,
                                                Annee_exercice=Annee_exercice,
                                                Service=service,
                                                Especes_des_unites=Especes_des_unites,
                                                Quantite=Quantite,
                                                Prix_unite=Prix_unite,
                                                Montant=Montant,
                                                Observations=Observations)

        return redirect(reverse('attestation', args=[service_id]))
    return render(request, "patrimoine/ajout_attestation.html", {"service": service})


# fonction d'ajout d'un ordre de sortie
@login_required
def ajout_ordre_sortie(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Budget = request.POST.get('Budget')
        imputation_administrative = request.POST.get('imputation_administrative')
        Designation_du_chapitre = request.POST.get('Designation_du_chapitre')
        libelle_article = request.POST.get('libelle_article')
        Subdivision_du_chapitre = request.POST.get('Subdivision_du_chapitre')
        Numero_du_journal = request.POST.get('Numero_du_journal')
        Nom_et_qualite_du_depositaire_comptable = request.POST.get('Nom_et_qualite_du_depositaire_comptable')
        Annee_exercice = request.POST.get("Annee_exercice")
        Numero_folio_du_grandlivre = request.POST.get("Numero_folio_du_grandlivre")
        Nomenclature = request.POST.get("Nomenclature")
        Designation_des_matieres_et_objets = request.POST.get("Designation_des_matieres_et_objets")
        Especes_des_unites = request.POST.get("Especes_des_unites")
        Quantites = request.POST.get("Quantites")
        Prix_unite = request.POST.get("Prix_unite")
        Valeurs_partielles = request.POST.get("Valeurs_partielles")
        Valeurs_par_numero_nomenclature = request.POST.get("Valeurs_par_numero_nomenclature")
        Numero_piece_justificative_sortie_correspondante = request.POST.get(
            "Numero_piece_justificative_sortie_correspondante")

        OrdreSortie.objects.create(Annee_exercice=Annee_exercice,
                                   Service=service,
                                   Budget=Budget,
                                   imputation_administrative=imputation_administrative,
                                   Designation_du_chapitre=Designation_du_chapitre,
                                   libelle_article=libelle_article,
                                   Subdivision_du_chapitre=Subdivision_du_chapitre,
                                   Numero_du_journal=Numero_du_journal,
                                   Nom_et_qualite_du_depositaire_comptable=Nom_et_qualite_du_depositaire_comptable,
                                   Numero_folio_du_grandlivre=Numero_folio_du_grandlivre,
                                   Nomenclature=Nomenclature,
                                   Designation_des_matieres_et_objets=Designation_des_matieres_et_objets,
                                   Especes_des_unites=Especes_des_unites,
                                   Quantites=Quantites,
                                   Prix_unite=Prix_unite,
                                   Valeurs_partielles=Valeurs_partielles,
                                   Valeurs_par_numero_nomenclature=Valeurs_par_numero_nomenclature,
                                   Numero_piece_justificative_sortie_correspondante=Numero_piece_justificative_sortie_correspondante)

        return redirect(reverse('ordre_sortie', args=[service_id]))
    return render(request, "patrimoine/ajout_ordre_sortie.html", {"service": service})


def ajout_pv_evaluation(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == "POST":
        Budget = request.POST.get('Budget'),
        imputation_administrative = request.POST.get('imputation_administrative')
        Designation_du_chapitre = request.POST.get('Designation_du_chapitre')
        libelle_article = request.POST.get('libelle_article')
        President_commission_recensement = request.POST.get('President_commission_recensement')
        Membres = request.POST.get('Membres')
        Nom_et_qualite_du_depositaire_comptable = request.POST.get('Nom_et_qualite_du_depositaire_comptable')
        Annee_exercice = request.POST.get('exercice')
        Nomenclature = request.POST.get('Nomenclature')
        Numero_ordre = request.POST.get('Numero_ordre')
        Designations_des_articles = request.POST.get('Designations_des_articles')
        Especes_des_unites = request.POST.get('Especes_des_unites')
        Prix_unitaire = request.POST.get('Prix_unitaire')
        Quantite = request.POST.get('Quantite')
        Montant = request.POST.get('Montant')
        Observations = request.POST.get('Observations')

        PvEvaluation.objects.create(Budget=Budget,
                                    Service=service,
                                    imputation_administrative=imputation_administrative,
                                    Designation_du_chapitre=Designation_du_chapitre,
                                    libelle_article=libelle_article,
                                    President_commission_recensement=President_commission_recensement,
                                    Membres=Membres,
                                    Nom_et_qualite_du_depositaire_comptable=Nom_et_qualite_du_depositaire_comptable,
                                    Annee_exercice=Annee_exercice,
                                    Nomenclature=Nomenclature,
                                    Numero_ordre=Numero_ordre,
                                    Designations_des_articles=Designations_des_articles,
                                    Especes_des_unites=Especes_des_unites,
                                    Prix_unitaire=Prix_unitaire,
                                    Quantite=Quantite,
                                    Montant=Montant,
                                    Observations=Observations)
        return redirect(reverse('pv_evaluation', args=[service_id]))
    return render(request, "patrimoine/ajout_pvevaluation.html", {'service':service})


def pv_evaluation(request, service_id):
    # Récupérer tous les objets PVRecensement
    service = get_object_or_404(Service, id=service_id)
    pvevaluations = PvEvaluation.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        pvevaluations = pvevaluations.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        pvevaluations = pvevaluations.filter(Annee_exercice=selected_annee_exercice)
    paginator = Paginator(pvevaluations, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Calculer les récapitulatifs en fonction des filtres appliqués
    recap = pvevaluations.values('Nomenclature').annotate(
        total_Montant=Sum('Montant'),
    ).order_by('Nomenclature')

    # Calculer le total général en fonction des données filtrées
    Total_general = {
        'total_Montant': recap.aggregate(Sum('total_Montant'))['total_Montant__sum'] or 0,
    }

    # Renvoyer les données au template
    context = {
        "pvevaluations": pvevaluations,
        "Total_general": Total_general,
        "recap": recap,
        "selected_nomenclature": selected_nomenclature,
        "selected_annee_exercice": selected_annee_exercice,
        "service": service,
        "page_obj": page_obj,
    }

    return render(request, "patrimoine/pv_evaluation.html", context)


# fonction d'affichage du procès verbal de recensement
@login_required
def PVRecensement_table(request, service_id):
    # Récupérer tous les objets PVRecensement
    service = get_object_or_404(Service, id=service_id)
    pvrecensements = PVRecensement.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        pvrecensements = pvrecensements.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        pvrecensements = pvrecensements.filter(Annee_exercice=selected_annee_exercice)
    paginator = Paginator(pvrecensements, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Calculer les récapitulatifs en fonction des filtres appliqués
    recap = pvrecensements.values('Nomenclature').annotate(
        total_existants=Sum('valeurs_des_existants'),
        total_excedents=Sum('valeurs_excedents_par_article'),
        total_deficits=Sum('valeurs_deficits_par_article')
    ).order_by('Nomenclature')

    # Calculer le total général en fonction des données filtrées
    Total_general = {
        'total_valeurs_des_existants': recap.aggregate(Sum('total_existants'))['total_existants__sum'] or 0,
        'total_valeurs_excedents_par_article': recap.aggregate(Sum('total_excedents'))['total_excedents__sum'] or 0,
        'total_valeurs_deficits_par_article': recap.aggregate(Sum('total_deficits'))['total_deficits__sum'] or 0,
    }

    # Renvoyer les données au template
    context = {
        "pvrecensements": pvrecensements,
        "Total_general": Total_general,
        "recap": recap,
        "selected_nomenclature": selected_nomenclature,
        "selected_annee_exercice": selected_annee_exercice,
        "service": service,
        "page_obj": page_obj,
    }

    return render(request, "patrimoine/pvrecensement.html", context)


# fonction d'affichage de l'inventaire
@login_required
def Inventaire_table(request, service_id):
    # Récupérer tous les objets Inventaire
    service = get_object_or_404(Service, id=service_id)
    inventaires = Inventaire.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        inventaires = inventaires.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        inventaires = inventaires.filter(Annee_exercice=selected_annee_exercice)

    paginator = Paginator(inventaires, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Calculer les récapitulatifs en fonction des données filtrées
    recap = inventaires.values('Nomenclature').annotate(
        decompte=Sum('Decompte'),
    ).order_by('Nomenclature')

    # Calculer le total général en fonction des données filtrées
    Total_general = {
        'total_decompte': recap.aggregate(Sum('decompte'))['decompte__sum'] or 0,
    }

    # Renvoyer les données au template
    context = {
        "inventaires": inventaires,
        "recap": recap,
        "Total_general": Total_general,
        "selected_nomenclature": selected_nomenclature,
        "selected_annee_exercice": selected_annee_exercice,
        "service": service,
        "page_obj": page_obj,
    }

    return render(request, "patrimoine/inventaire.html", context)


@login_required
def recapitulatif_inventaire(request, service_id):
    service = get_object_or_404(Service, id=service_id)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Récapitulatif des données d'inventaire
    inventaire_recap = Inventaire.objects.filter(Service=service)
    if selected_nomenclature:
        inventaire_recap = inventaire_recap.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        inventaire_recap = inventaire_recap.filter(Annee_exercice=selected_annee_exercice)

    inventaire_recap = (
        inventaire_recap
        .values('Nomenclature', 'Annee_exercice')
        .annotate(
            total_valeur=Sum('Decompte'),
            total_prix_janvier=Sum(F('Quantite_existant_1er_janvier') * F('Prix_de_l_unite')),
            total_existant_et_entrees=Sum(
                (F('Quantite_existant_1er_janvier') + F('Quantite_entree_pendant_l_annee')) * F('Prix_de_l_unite')
            ),
            total_reste=Sum(F('Quantite_reste_31_decembre') * F('Prix_de_l_unite'))
        )
        .order_by('Nomenclature')
    )

    # Récapitulatif des données d'état appréciatif
    etatappreciatif_recap = EtatApreciatif.objects.filter(Service=service)
    if selected_nomenclature:
        etatappreciatif_recap = etatappreciatif_recap.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        etatappreciatif_recap = etatappreciatif_recap.filter(Annee_exercice=selected_annee_exercice)

    etatappreciatif_recap = (
        etatappreciatif_recap
        .values('Nomenclature', 'Annee_exercice')
        .annotate(
            total_entrees_annee=Sum('Charge'),  # Correspond à Charge
            total_sorties=Sum('Decharge')  # Correspond à Decharge
        )
        .order_by('Nomenclature')
    )

    # Fusionner les récapitulatifs
    recapitulatif = []
    for inv_item in inventaire_recap:
        # Chercher la correspondance dans les données d'état appréciatif
        matching_etat = next((etat for etat in etatappreciatif_recap if
                              etat['Nomenclature'] == inv_item['Nomenclature'] and etat['Annee_exercice'] == inv_item[
                                  'Annee_exercice']), None)

        recapitulatif.append({
            'Nomenclature': inv_item['Nomenclature'],
            'Annee_exercice': inv_item['Annee_exercice'],
            'total_valeur': inv_item['total_valeur'],
            'total_prix_janvier': inv_item['total_prix_janvier'],
            'total_existant_et_entrees': inv_item['total_existant_et_entrees'],
            'total_reste': inv_item['total_reste'],
            'total_entrees_annee': matching_etat['total_entrees_annee'] if matching_etat else 0,
            'total_sorties': matching_etat['total_sorties'] if matching_etat else 0,
        })

    # Calculer le total général
    total_general = {
        'total_nomenclature': sum(item['total_valeur'] for item in recapitulatif),
        'total_prix_janvier': sum(item['total_prix_janvier'] for item in recapitulatif),
        'total_entrees_annee': sum(item['total_entrees_annee'] for item in recapitulatif),
        'total_existant_et_entrees': sum(item['total_existant_et_entrees'] for item in recapitulatif),
        'total_sorties': sum(item['total_sorties'] for item in recapitulatif),
        'total_reste': sum(item['total_reste'] for item in recapitulatif)
    }

    # Préparer le contexte
    context = {
        'recapitulatif': recapitulatif,
        'total_general': total_general,
        'service': service,
        'etatappreciatif': etatappreciatif_recap,
        'selected_nomenclature': selected_nomenclature,
        'selected_annee_exercice': selected_annee_exercice,
    }
    return render(request, 'patrimoine/recapitulatif.html', context)


# fonction d'affichage de l'etat appreciatif
@login_required
def EtatApreciatif_table(request, service_id):
    # Récupérer tous les objets EtatApreciatif
    service = get_object_or_404(Service, id=service_id)
    etatappreciatifs = EtatApreciatif.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        etatappreciatifs = etatappreciatifs.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        etatappreciatifs = etatappreciatifs.filter(Annee_exercice=selected_annee_exercice)

    paginator = Paginator(etatappreciatifs, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Calculer les récapitulatifs en fonction des données filtrées
    recap = etatappreciatifs.values('Nomenclature').annotate(
        charge=Sum('Charge'),
        decharge=Sum('Decharge'),
    ).order_by('Nomenclature')

    # Calculer le total général en fonction des données filtrées
    Total_general = {
        'total_charge': recap.aggregate(Sum('charge'))['charge__sum'] or 0,
        'total_decharge': recap.aggregate(Sum('decharge'))['decharge__sum'] or 0,
    }

    # Renvoyer les données au template
    context = {
        "etatappreciatifs": etatappreciatifs,
        "recap": recap,
        "Total_general": Total_general,
        "selected_nomenclature": selected_nomenclature,
        "selected_annee_exercice": selected_annee_exercice,
        "service": service,
        'page_obj': page_obj,
    }

    return render(request, "patrimoine/etat_appreciatif.html", context)


# fonction d'affichage de l'ordre entree
@login_required
def OrdreEntree_table(request, service_id):
    # Récupérer tous les objets OrdreEntree
    service = get_object_or_404(Service, id=service_id)
    ordreEntrees = OrdreEntree.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        ordreEntrees = ordreEntrees.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        ordreEntrees = ordreEntrees.filter(Annee_exercice=selected_annee_exercice)

    paginator = Paginator(ordreEntrees, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Calculer les récapitulatifs en fonction des données filtrées
    recap = ordreEntrees.values('Nomenclature').annotate(
        Valeurs_par_numero_nomenclature=Sum('Valeurs_par_numero_nomenclature')
    ).order_by('Nomenclature')

    # Calculer le total général en fonction des données filtrées
    Total_general = {
        'total_valeurs_par_nomenclature': recap.aggregate(Sum('Valeurs_par_numero_nomenclature'))[
                                              'Valeurs_par_numero_nomenclature__sum'] or 0,
    }

    # Renvoyer les données au template
    context = {
        "ordreEntrees": ordreEntrees,
        "recap": recap,
        "Total_general": Total_general,
        "selected_nomenclature": selected_nomenclature,
        "selected_annee_exercice": selected_annee_exercice,
        "service": service,
        "page_obj": page_obj,
    }

    return render(request, "patrimoine/ordre_entree.html", context)


# fonction d'affichage de l'ordre de sortie
@login_required
def OrdreSortie_table(request, service_id):
    # Récupérer tous les objets OrdreSortie
    service = get_object_or_404(Service, id=service_id)
    ordresorties = OrdreSortie.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        ordresorties = ordresorties.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        ordresorties = ordresorties.filter(Annee_exercice=selected_annee_exercice)

    paginator = Paginator(ordresorties, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Calculer les récapitulatifs en fonction des données filtrées
    recap = ordresorties.values('Nomenclature').annotate(
        Valeurs_par_numero_nomenclature=Sum('Valeurs_par_numero_nomenclature')
    ).order_by('Nomenclature')

    # Calculer le total général en fonction des données filtrées
    Total_general = {
        'total_valeurs_par_nomenclature': recap.aggregate(Sum('Valeurs_par_numero_nomenclature'))[
                                              'Valeurs_par_numero_nomenclature__sum'] or 0,
    }

    # Renvoyer les données au template
    context = {
        "ordresorties": ordresorties,
        "Total_general": Total_general,
        "recap": recap,
        "selected_nomenclature": selected_nomenclature,
        "selected_annee_exercice": selected_annee_exercice,
        "service": service,
        "page_obj": page_obj
    }

    return render(request, "patrimoine/ordre_sortie.html", context)


# fonction d'affichage de l'attestation de prise en charge
@login_required
def AttestationPriseEnCharge_table(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    attestationpriseEncharges = AttestationPriseEnCharge.objects.filter(Service=service)

    # Récupérer les valeurs des filtres depuis les paramètres GET
    selected_nomenclature = request.GET.get('nomenclature')
    selected_annee_exercice = request.GET.get('annee_exercice')

    # Filtrer les données en fonction des filtres sélectionnés
    if selected_nomenclature:
        attestationpriseEncharges = attestationpriseEncharges.filter(Nomenclature=selected_nomenclature)
    if selected_annee_exercice:
        attestationpriseEncharges = attestationpriseEncharges.filter(Annee_exercice=selected_annee_exercice)

    paginator = Paginator(attestationpriseEncharges, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    recap = AttestationPriseEnCharge.objects.values('Designation_des_matieres_et_objets').annotate(
        total_general=Sum('Montant'),
        ).order_by('Designation_des_matieres_et_objets')
    Total_general = {
        'total_general': recap.aggregate(Sum('total_general'))['total_general__sum'] or 0,
    }
    return render(request, "patrimoine/attestation.html",
                  {"attestationpriseEncharges": attestationpriseEncharges, "Total_general": Total_general,
                   "recap": recap, "service": service})


# mettre le mot nomenclature dans un dictionnaire pour eviter le problème dans le filtre
FIELD_MAPPING = {
    'nomenclature': 'Nomenclature',
    'annee_exercice': 'Annee_exercice',
    # Ajoutez d'autres correspondances si nécessaire
}


def export_data(request, model_class, fields, filename_prefix, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()  # Normaliser en minuscules

    # Debugging: Afficher les filtres et le format
    print(f"Filtres : {filters}")
    print(f"Format demandé : {format_export}")

    # Récupérer le service spécifique par son ID
    service = get_object_or_404(Service, id=service_id)

    # Récupérer toutes les données pour le service spécifié
    data_queryset = model_class.objects.filter(Service=service)

    # Appliquer d'autres filtres
    for key, value in filters.items():
        if value and value.lower() != 'none':  # Ignorez 'None'
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(model_class, correct_key):
                print(f"Filtrage avec : {{{correct_key}: {value}}}")  # Log de filtrage
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    # Exporter selon le format choisi
    if format_export == 'pdf':
        return export_to_pdf(data_queryset, fields, filename_prefix)
    elif format_export == 'excel':
        return export_to_excel(data_queryset, fields, filename_prefix)
    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)


def export_to_pdf(data_queryset, fields, filename_prefix):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.pdf"'
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    # Préparer les éléments du PDF
    elements = []

    # Ajouter en-tête
    title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
    elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("************", title_style))
    elements.append(Spacer(1, 20))

    # Ajouter les informations de budget (à remplacer par les valeurs réelles)
    budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)
    elements.append(Paragraph("BUDGET: ", budget_style))
    elements.append(Paragraph("EXERCICE: ", budget_style))
    elements.append(Paragraph("IMPUTATION ADMINISTRATIVE: ", budget_style))
    elements.append(Paragraph("CHAPITRE:  - ARTICLE: - PARAGRAPHE: ", budget_style))
    elements.append(Spacer(1, 20))

    # Préparer les données pour le tableau
    data = [fields]  # En-tête du tableau

    # Ajouter les données des objets filtrés
    for item in data_queryset:
        row = [getattr(item, field) for field in fields]
        data.append(row)

    # Créer le tableau PDF
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Correction pour dessiner les bordures du tableau
    ]))

    # Ajouter le tableau au document PDF
    elements.append(table)
    doc.build(elements)

    # Récupérer le contenu du PDF
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def export_to_excel(data_queryset, fields, filename_prefix):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"

    # En-têtes personnalisées
    ws.append(['REPOBLIKAN\'I MADAGASIKARA', '', '', '', '', '', '', '', '', '', '', '', ''])
    ws.append(['Fitiavana-Tanindrazana-Fandrosoana', '', '', '', '', '', '', '', '', '', '', '', ''])
    ws.append(['************', '', '', '', '', '', '', '', '', '', '', '', ''])

    # Ajouter les informations de budget (à remplacer par les valeurs réelles)
    ws.append(['BUDGET', '[Budget]'])
    ws.append(['EXERCICE', '[Exercice]'])
    ws.append(['IMPUTATION ADMINISTRATIVE', '[Imputation]'])
    ws.append(['CHAPITRE', '[Chapitre]', 'ARTICLE', '[Article]', 'PARAGRAPHE', '[Paragraphe]'])
    ws.append([''])  # Ajouter une ligne vide

    # En-têtes du tableau
    ws.append(fields)

    # Ajouter les données
    for item in data_queryset:
        row = [getattr(item, field) for field in fields]
        ws.append(row)

    # Enregistrer le fichier Excel dans la réponse
    wb.save(response)
    return response


def export_pvrecensement(request, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = PVRecensement.objects.filter(Service=service)
    fields = [
        'Annee_exercice', 'Nomenclature', 'Designation_materiels', 'Especes_unites',
        'Prix_unites', 'Quantites_d_apres_ecriture', 'Quantites_par_recensement',
        'Quantites_execedent_par_article', 'Quantites_deficient_par_article',
        'valeurs_excedents_par_article', 'valeurs_excedents_par_nomenclature',
        'valeurs_deficits_par_article', 'valeurs_des_deficits_par_nomenclature',
        'valeurs_des_existants', 'Observations'
    ]

    field_headers = [field.replace('_', ' ').title() for field in fields]

    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(PVRecensement, correct_key):
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    # Initialisation des variables pour éviter l'erreur
    budget = ""
    imputation_administrative = ""
    chapitre = ""
    article = ""
    magasin_service = ""
    recenseur = ""
    depositaire = ""

    # Récupérer les données supplémentaires si data_queryset contient des données
    if data_queryset.exists():
        pv_item = data_queryset.first()
        budget = pv_item.Budget or ""
        imputation_administrative = pv_item.Imputation_administrative or ""
        chapitre = pv_item.Designation_chapitre or ""
        article = pv_item.Libelle_article or ""
        magasin_service = pv_item.Designation_magasin_ou_materiels_service or ""
        recenseur = pv_item.Nom_et_qualite_recenseur or ""
        depositaire = pv_item.Nom_et_qualite_depositaire_comptable or ""

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="PVRecensement.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))

        elements = []

        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)
        elements.append(Paragraph(f"BUDGET: {budget}", budget_style))
        elements.append(Paragraph(f"IMPUTATION ADMINISTRATIVE: {imputation_administrative}", budget_style))
        elements.append(Paragraph(f"CHAPITRE: {chapitre} - ARTICLE: {article} - PARAGRAPHE: ", budget_style))
        elements.append(Paragraph(f"DESIGNATION MAGASIN/SERVICE: {magasin_service}", budget_style))
        elements.append(Paragraph(f"NOM ET QUALITE DU RECENSEUR: {recenseur}", budget_style))
        elements.append(Paragraph(f"NOM ET QUALITE DU DEPOSITAIRE: {depositaire}", budget_style))
        elements.append(Spacer(1, 20))

        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1, wordWrap='CJK')
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field)
                if isinstance(value, str):
                    paragraph = Paragraph(value, ParagraphStyle('Normal', fontName='Helvetica', fontSize=10, wordWrap='CJK'))
                    row.append(paragraph)
                else:
                    row.append(value)
            data.append(row)

        column_widths = [50, 50, 100, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50]
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="PVRecensement.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"

        ws.append(['REPOBLIKAN\'I MADAGASIKARA'])
        ws.append(['Fitiavana-Tanindrazana-Fandrosoana'])
        ws.append(['************'])
        ws.append(['BUDGET', budget])
        ws.append(['IMPUTATION ADMINISTRATIVE', imputation_administrative])
        ws.append(['CHAPITRE', chapitre, 'ARTICLE', article])
        ws.append(['MAGASIN/SERVICE', magasin_service])
        ws.append(['RECENSEUR', recenseur])
        ws.append(['DEPOSITAIRE', depositaire])
        ws.append([''])

        ws.append(field_headers)

        for item in data_queryset:
            row = [getattr(item, field) for field in fields]
            ws.append(row)

        wb.save(response)
        return response
    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)

# Fonction pour l'exportation du données de la table Etat Apreciatif
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import ParagraphStyle
from openpyxl import Workbook
import io

def export_etat_apreciatif(request, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = EtatApreciatif.objects.filter(Service=service)

    # Champs à utiliser dans le tableau
    fields = [
        'Annee_exercice', 'Nomenclature', 'Numero_du_piece_justificative',
        'Date_du_piece_justificative', 'Designations_sommaire_des_operations',
        'Charge', 'Decharge'
    ]

    # En-têtes pour le tableau (sans les nouveaux champs)
    field_headers = [field.replace('_', ' ').title() for field in fields]

    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(EtatApreciatif, correct_key):
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="PVRecensement.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))  # Mode paysage

        elements = []

        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        # Ajouter les champs supplémentaires dans l'en-tête
        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)
        elements.append(
            Paragraph(f"BUDGET: {data_queryset.first().Budget if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(
            Paragraph(f"MATERIEL EN APPROVISIONNEMENT OU EN SERVICE: {data_queryset.first().Materiel_en_approvisionnement_ou_en_service if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(
            Paragraph(f"DÉSIGNATION DU CHAPITRE: {data_queryset.first().Designation_du_chapitre if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(
            Paragraph(f"LIBELLE ARTICLE: {data_queryset.first().libelle_article if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(
            Paragraph(f"SUBDIVISION DU MAGASIN OU DE LA CATEGORIE DU MATERIEL EN SERVICE: {data_queryset.first().Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(
            Paragraph(f"NOMBRE EN TOUTES LETTRES: {data_queryset.first().Nombre_en_toutes_lettres if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(
            Paragraph(f"NOM ET QUALITE DU DEPOSITAIRE COMPTABLE: {data_queryset.first().Nom_et_qualite_du_depositaire_comptable if data_queryset.exists() else 'N/A'}", budget_style))
        elements.append(Spacer(1, 20))

        # Style de cellule pour le tableau
        cell_style = ParagraphStyle(name='CellStyle', fontName='Helvetica', fontSize=10, wordWrap='CJK')

        # Préparer les en-têtes avec saut de ligne
        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1)
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        # Ajouter les données des objets filtrés
        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field)
                if isinstance(value, str):
                    paragraph = Paragraph(value, cell_style)  # Utiliser le style de cellule
                    row.append(paragraph)
                else:
                    row.append(value)
            data.append(row)

        column_widths = [100] * len(fields)  # Ajustez si nécessaire

        # Créer le tableau PDF avec les en-têtes ajustés
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)

        # Construire le document PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        # Création d'un fichier Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "État Appréciatif"

        # En-têtes pour l'Excel
        headers = field_headers.copy()
        headers.extend([
            'Budget', 'Materiel_en_approvisionnement_ou_en_service',
            'Designation_du_chapitre', 'libelle_article',
            'Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service',
            'Nombre_en_toutes_lettres', 'Nom_et_qualite_du_depositaire_comptable'
        ])

        # Écrire les en-têtes dans la feuille Excel
        worksheet.append(headers)

        # Écrire les données dans la feuille Excel
        for item in data_queryset:
            row = [getattr(item, field) for field in fields]
            row.extend([
                item.Budget, item.Materiel_en_approvisionnement_ou_en_service,
                item.Designation_du_chapitre, item.libelle_article,
                item.Subdivision_du_magasin_ou_de_la_categorie_du_materiel_en_service,
                item.Nombre_en_toutes_lettres, item.Nom_et_qualite_du_depositaire_comptable
            ])
            worksheet.append(row)

        # Utiliser BytesIO pour le téléchargement
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Configuration de la réponse pour le téléchargement
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="etat_apreciatif.xlsx"'
        return response

    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)

# Fonction pour l'exportation des données de la table Inventaire
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import ParagraphStyle
from openpyxl import Workbook
import io


def export_inventaire(request, service_id):
    # Récupérer les filtres et le format d'export
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = Inventaire.objects.filter(Service=service)
    fields = [
        'Annee_exercice', 'Nomenclature', 'Numero_folio_grand_livre',
        'Designation_materiels', 'Especes_des_unites', 'Prix_de_l_unite',
        'Quantite_existant_1er_janvier', 'Quantite_entree_pendant_l_annee',
        'Quantite_sortie_pendant_l_annee', 'Quantite_reste_31_decembre',
        'Decompte', 'Observation'
    ]

    # Mapping des en-têtes sans les underscores pour les afficher en plusieurs lignes
    field_headers = [field.replace('_', ' ').title() for field in fields]

    # Filtrer les données selon les paramètres fournis
    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(Inventaire, correct_key):  # Changer PVRecensement en Inventaire
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    # Initialiser les valeurs
    budget_value = ''
    exercice_value = ''  # Ajouter votre logique pour l'exercice ici
    imputation_value = ''
    chapitre_value = ''
    article_value = ''
    paragraphe_value = ''

    if data_queryset.exists():
        first_item = data_queryset.first()
        budget_value = first_item.Budget if hasattr(first_item, 'Budget') else ''
        exercice_value = first_item.Annee_exercice if hasattr(first_item, 'Annee_exercice') else ''
        imputation_value = first_item.imputation_administrative if hasattr(first_item,
                                                                           'imputation_administrative') else ''
        chapitre_value = first_item.Designation_du_chapitre if hasattr(first_item, 'Designation_du_chapitre') else ''
        article_value = first_item.libelle_article if hasattr(first_item, 'libelle_article') else ''
        paragraphe_value = first_item.Nom_et_qualite_du_recenseur if hasattr(first_item,
                                                                             'Nom_et_qualite_du_recenseur') else ''

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Inventaire.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))  # Mode paysage

        elements = []

        # Styles
        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        # Ajouter les informations de budget
        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)
        elements.append(Paragraph(f"BUDGET: {budget_value}", budget_style))
        elements.append(Paragraph(f"EXERCICE: {exercice_value}", budget_style))
        elements.append(Paragraph(f"IMPUTATION ADMINISTRATIVE: {imputation_value}", budget_style))
        elements.append(
            Paragraph(f"CHAPITRE: {chapitre_value} - ARTICLE: {article_value} - PARAGRAPHE: {paragraphe_value}",
                      budget_style))
        elements.append(Spacer(1, 20))

        # Préparer les en-têtes avec saut de ligne
        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1)
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        # Ajouter les données des objets filtrés
        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field, '')
                paragraph = Paragraph(value if isinstance(value, str) else str(value),
                                      ParagraphStyle('Normal', fontName='Helvetica', fontSize=10, wordWrap='CJK'))
                row.append(paragraph)
            data.append(row)

        column_widths = [50, 50, 50, 100, 50, 50, 70, 70, 70, 70, 50, 80]

        # Créer le tableau PDF
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('WORDWRAP', (0, 1), (-1, -1), True),
        ]))

        # Ajouter la table au document PDF
        elements.append(table)

        # Construire le document PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Inventaire.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"

        # Ajouter les en-têtes
        ws.append(['REPOBLIKAN\'I MADAGASIKARA', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['Fitiavana-Tanindrazana-Fandrosoana', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['************', '', '', '', '', '', '', '', '', '', '', '', ''])

        # Ajouter les informations de budget
        ws.append(['BUDGET', budget_value])
        ws.append(['EXERCICE', exercice_value])
        ws.append(['IMPUTATION ADMINISTRATIVE', imputation_value])
        ws.append(['CHAPITRE', chapitre_value, 'ARTICLE', article_value, 'PARAGRAPHE', paragraphe_value])
        ws.append([''])  # Ligne vide

        ws.append(field_headers)  # En-têtes complètes

        # Ajouter les données
        for item in data_queryset:
            row = [getattr(item, field, '') for field in fields]
            ws.append(row)

        wb.save(response)
        return response
    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)


# fonction pour l'exportation du données de la table Ordre entrée
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import ParagraphStyle
from openpyxl import Workbook
import io

# Assurez-vous d'importer vos modèles
from .models import OrdreEntree, Service

def export_ordre_entree(request, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = OrdreEntree.objects.filter(Service=service)
    fields = [
        'Annee_exercice', 'Numero_folio_du_grandlivre', 'Nomenclature',
        'Designation_des_matieres_et_objets', 'Especes_des_unites', 'Quantites',
        'Prix_unite', 'Valeurs_partielles', 'Valeurs_par_numero_nomenclature',
        'Numero_piece_justificative_sortie_correspondante'
    ]

    # Mapping des en-têtes sans les underscores pour les afficher en plusieurs lignes
    field_headers = [field.replace('_', ' ').title() for field in fields]

    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(OrdreEntree, correct_key):
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    # Initialiser les valeurs pour éviter UnboundLocalError
    budget_value = ''
    exercice_value = ''  # Ajoutez votre logique pour l'exercice ici
    imputation_value = ''
    chapitre_value = ''
    article_value = ''
    subdivision_value = ''
    numero_journal_value = ''
    paragraphe_value = ''

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Ordre_entree.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))  # Mode paysage

        elements = []

        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)

        # Récupérer les valeurs si des données existent
        if data_queryset.exists():
            first_item = data_queryset.first()
            budget_value = first_item.Budget or ''
            exercice_value = first_item.Annee_exercice or ''  # Ajustez selon vos besoins
            imputation_value = first_item.imputation_administrative or ''
            chapitre_value = first_item.Designation_du_chapitre or ''
            article_value = first_item.libelle_article or ''
            subdivision_value = first_item.Subdivision_du_chapitre or ''
            numero_journal_value = first_item.Numero_du_journal or ''
            paragraphe_value = first_item.Nom_et_qualite_du_depositaire_comptable or ''

        # Ajout des informations de budget dans le PDF
        elements.append(Paragraph(f"BUDGET: {budget_value}", budget_style))
        elements.append(Paragraph(f"EXERCICE: {exercice_value}", budget_style))
        elements.append(Paragraph(f"IMPUTATION ADMINISTRATIVE: {imputation_value}", budget_style))
        elements.append(Paragraph(f"CHAPITRE: {chapitre_value} - ARTICLE: {article_value} - PARAGRAPHE: {paragraphe_value}", budget_style))
        elements.append(Paragraph(f"SUBDIVISION DU CHAPITRE: {subdivision_value}", budget_style))
        elements.append(Paragraph(f"NUMERO DU JOURNAL: {numero_journal_value}", budget_style))
        elements.append(Spacer(1, 20))

        # Préparer les en-têtes avec saut de ligne
        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1)
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        # Ajouter les données des objets filtrés
        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field)
                if isinstance(value, str):
                    paragraph = Paragraph(value, ParagraphStyle('Normal', fontName='Helvetica', fontSize=10))
                    row.append(paragraph)
                else:
                    row.append(value)
            data.append(row)

        column_widths = [50, 80, 70, 120, 60, 60, 60, 60, 60, 100]

        # Créer le tableau PDF avec les en-têtes ajustés
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)

        # Construire le document PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Ordre_entree.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordre Entree"

        # Ajoutez les informations de budget uniquement si data_queryset contient des éléments
        if data_queryset.exists():
            first_item = data_queryset.first()
            budget_value = first_item.Budget or ''
            exercice_value = first_item.Annee_exercice or ''  # Ajustez cette ligne selon vos besoins
            imputation_value = first_item.imputation_administrative or ''
            chapitre_value = first_item.Designation_du_chapitre or ''
            article_value = first_item.libelle_article or ''
            subdivision_value = first_item.Subdivision_du_chapitre or ''
            numero_journal_value = first_item.Numero_du_journal or ''
            paragraphe_value = first_item.Nom_et_qualite_du_depositaire_comptable or ''

        # Ajout des informations de budget dans Excel
        ws.append(['BUDGET', budget_value])
        ws.append(['EXERCICE', exercice_value])
        ws.append(['IMPUTATION ADMINISTRATIVE', imputation_value])
        ws.append(['CHAPITRE', chapitre_value, 'ARTICLE', article_value, 'PARAGRAPHE', paragraphe_value])
        ws.append(['SUBDIVISION DU CHAPITRE', subdivision_value])
        ws.append(['NUMERO DU JOURNAL', numero_journal_value])
        ws.append([''])  # Ligne vide

        ws.append(field_headers)  # En-têtes complètes

        for item in data_queryset:
            row = [getattr(item, field) for field in fields]
            ws.append(row)

        wb.save(response)
        return response

    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)


# fonction pour l'exportation du données de la table Attestation prise en charge

def export_attestation_prise_en_charge(request, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = AttestationPriseEnCharge.objects.filter(Service=service)
    fields = [
        'Designation_des_matieres_et_objets', 'Especes_des_unites',
        'Quantite', 'Prix_unite', 'Montant', 'Observations'
    ]

    # Mapping des en-têtes sans les underscores pour les afficher en plusieurs lignes
    field_headers = [field.replace('_', ' ').title() for field in fields]

    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(AttestationPriseEnCharge, correct_key):
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Attestation_de_prise_en_charge.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))  # Mode paysage

        elements = []

        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)
        elements.append(Paragraph("BUDGET: ", budget_style))
        elements.append(Paragraph("EXERCICE: ", budget_style))
        elements.append(Paragraph("IMPUTATION ADMINISTRATIVE: ", budget_style))
        elements.append(Paragraph("CHAPITRE:  - ARTICLE: - PARAGRAPHE: ", budget_style))
        elements.append(Spacer(1, 20))

        # Préparer les en-têtes avec saut de ligne
        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1)
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        # Ajouter les données des objets filtrés
        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field)
                if isinstance(value, str):
                    paragraph = Paragraph(value, ParagraphStyle('Normal', fontName='Helvetica', fontSize=10))
                    row.append(paragraph)
                else:
                    row.append(value)
            data.append(row)

        column_widths = [200, 80, 80, 80, 80, 120]

        # Créer le tableau PDF avec les en-têtes ajustés
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        def add_table_with_page_break(doc, table):
            height = sum([element.wrap(doc.width, doc.bottomMargin)[1] for element in elements])
            table_height = table.wrap(doc.width, doc.bottomMargin)[1]

            if height + table_height > doc.pagesize[1] - doc.topMargin - doc.bottomMargin:
                elements.append(PageBreak())  # Nouvelle page

            elements.append(table)

        add_table_with_page_break(doc, table)

        # Construire le document PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Attestation_de_prise_en_charge.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Attestation"

        ws.append(['REPOBLIKAN\'I MADAGASIKARA', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['Fitiavana-Tanindrazana-Fandrosoana', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['************', '', '', '', '', '', '', '', '', '', '', '', ''])

        ws.append(['BUDGET', '[Budget]'])
        ws.append(['EXERCICE', '[Exercice]'])
        ws.append(['IMPUTATION ADMINISTRATIVE', '[Imputation]'])
        ws.append(['CHAPITRE', '[Chapitre]', 'ARTICLE', '[Article]', 'PARAGRAPHE', '[Paragraphe]'])
        ws.append([''])  # Ligne vide

        ws.append(field_headers)  # En-têtes complètes

        for item in data_queryset:
            row = [getattr(item, field) for field in fields]
            ws.append(row)

        wb.save(response)
        return response

    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)


def export_ordre_sortie(request, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = OrdreSortie.objects.filter(Service=service)
    fields = [
        'Annee_exercice', 'Numero_folio_du_grandlivre', 'Nomenclature',
        'Designation_des_matieres_et_objets', 'Especes_des_unites', 'Quantites',
        'Prix_unite', 'Valeurs_partielles', 'Valeurs_par_numero_nomenclature',
        'Numero_piece_justificative_sortie_correspondante'
    ]

    # Mapping des en-têtes sans les underscores pour les afficher en plusieurs lignes
    field_headers = [field.replace('_', ' ').title() for field in fields]

    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(OrdreSortie, correct_key):
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    # Initialiser les valeurs pour éviter UnboundLocalError
    budget_value = ''
    exercice_value = ''
    imputation_value = ''
    chapitre_value = ''
    article_value = ''
    subdivision_value = ''
    numero_journal_value = ''
    paragraphe_value = ''

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Ordre_de_sortie.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))  # Mode paysage

        elements = []

        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)

        # Récupérer les valeurs si des données existent
        if data_queryset.exists():
            first_item = data_queryset.first()
            budget_value = first_item.Budget or ''
            exercice_value = first_item.Annee_exercice or ''
            imputation_value = first_item.imputation_administrative or ''
            chapitre_value = first_item.Designation_du_chapitre or ''
            article_value = first_item.libelle_article or ''
            subdivision_value = first_item.Subdivision_du_chapitre or ''
            numero_journal_value = first_item.Numero_du_journal or ''
            paragraphe_value = first_item.Nom_et_qualite_du_depositaire_comptable or ''

        # Ajout des informations de budget dans le PDF
        elements.append(Paragraph(f"BUDGET: {budget_value}", budget_style))
        elements.append(Paragraph(f"EXERCICE: {exercice_value}", budget_style))
        elements.append(Paragraph(f"IMPUTATION ADMINISTRATIVE: {imputation_value}", budget_style))
        elements.append(Paragraph(f"CHAPITRE: {chapitre_value} - ARTICLE: {article_value} - PARAGRAPHE: {paragraphe_value}", budget_style))
        elements.append(Paragraph(f"SUBDIVISION DU CHAPITRE: {subdivision_value}", budget_style))
        elements.append(Paragraph(f"NUMERO DU JOURNAL: {numero_journal_value}", budget_style))
        elements.append(Spacer(1, 20))

        # Préparer les en-têtes avec saut de ligne
        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1)
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        # Ajouter les données des objets filtrés
        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field)
                if isinstance(value, str):
                    paragraph = Paragraph(value, ParagraphStyle('Normal', fontName='Helvetica', fontSize=10))
                    row.append(paragraph)
                else:
                    row.append(value)
            data.append(row)

        column_widths = [60, 80, 100, 120, 70, 60, 60, 80, 80]

        # Créer le tableau PDF sans les en-têtes de budget
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)

        # Construire le document PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Ordre_de_sortie.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordre de Sortie"

        # Ajoutez les informations de budget uniquement si data_queryset contient des éléments
        if data_queryset.exists():
            first_item = data_queryset.first()
            budget_value = first_item.Budget or ''
            exercice_value = first_item.Annee_exercice or ''
            imputation_value = first_item.imputation_administrative or ''
            chapitre_value = first_item.Designation_du_chapitre or ''
            article_value = first_item.libelle_article or ''
            subdivision_value = first_item.Subdivision_du_chapitre or ''
            numero_journal_value = first_item.Numero_du_journal or ''
            paragraphe_value = first_item.Nom_et_qualite_du_depositaire_comptable or ''

        # Ajout des informations de budget dans Excel
        ws.append(['BUDGET', budget_value])
        ws.append(['EXERCICE', exercice_value])
        ws.append(['IMPUTATION ADMINISTRATIVE', imputation_value])
        ws.append(['CHAPITRE', chapitre_value, 'ARTICLE', article_value, 'PARAGRAPHE', paragraphe_value])
        ws.append(['SUBDIVISION DU CHAPITRE', subdivision_value])
        ws.append(['NUMERO DU JOURNAL', numero_journal_value])
        ws.append([''])  # Ligne vide

        ws.append(field_headers)  # En-têtes complètes

        for item in data_queryset:
            row = [getattr(item, field) for field in fields]
            ws.append(row)

        wb.save(response)
        return response

    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)


def export_pvevaluation(request, service_id):
    filters = {k: request.GET.get(k) for k in request.GET.keys() if k not in ['format']}
    format_export = request.GET.get('format', '').lower()

    service = get_object_or_404(Service, id=service_id)
    data_queryset = PvEvaluation.objects.filter(Service=service)
    fields = [
        'Annee_exercice', 'Nomenclature', 'Numero_ordre', 'Designations_des_articles',
        'Especes_des_unites', 'Prix_unitaire', 'Quantite',
        'Montant', 'Observations',
    ]

    # Mapping des en-têtes sans les underscores pour les afficher en plusieurs lignes
    field_headers = [field.replace('_', ' ').title() for field in fields]

    for key, value in filters.items():
        if value and value.lower() != 'none':
            correct_key = FIELD_MAPPING.get(key, key)
            if hasattr(PvEvaluation, correct_key):
                data_queryset = data_queryset.filter(**{correct_key: value})
            else:
                return HttpResponse(f"Le champ '{key}' n'existe pas dans le modèle.", status=400)

    if format_export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="PVEvaluation.pdf"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=(A4[1], A4[0]))  # Mode paysage

        elements = []

        title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=14, alignment=1)
        elements.append(Paragraph("REPOBLIKAN'I MADAGASIKARA", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fitiavana-Tanindrazana-Fandrosoana", title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("************", title_style))
        elements.append(Spacer(1, 20))

        budget_style = ParagraphStyle(name='BudgetStyle', fontName='Helvetica', fontSize=10, alignment=1)

        # Récupérer les valeurs si des données existent
        if data_queryset.exists():
            first_item = data_queryset.first()
            budget_value = first_item.Budget or ''
            exercice_value = first_item.Annee_exercice or ''
            imputation_value = first_item.imputation_administrative or ''
            chapitre_value = first_item.Designation_du_chapitre or ''
            article_value = first_item.libelle_article or ''
            paragraphe_value = first_item.Nom_et_qualite_du_depositaire_comptable or ''

            elements.append(Paragraph(f"BUDGET: {budget_value}", budget_style))
            elements.append(Paragraph(f"EXERCICE: {exercice_value}", budget_style))
            elements.append(Paragraph(f"IMPUTATION ADMINISTRATIVE: {imputation_value}", budget_style))
            elements.append(Paragraph(f"CHAPITRE: {chapitre_value} - ARTICLE: {article_value} - PARAGRAPHE: {paragraphe_value}", budget_style))
            elements.append(Spacer(1, 20))

        # Préparer les en-têtes avec saut de ligne
        header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=10, alignment=1)
        header_paragraphs = [Paragraph(header, header_style) for header in field_headers]
        data = [header_paragraphs]

        # Ajouter les données des objets filtrés
        for item in data_queryset:
            row = []
            for field in fields:
                value = getattr(item, field)
                if isinstance(value, str):
                    paragraph = Paragraph(value, ParagraphStyle('Normal', fontName='Helvetica', fontSize=10))
                    row.append(paragraph)
                else:
                    row.append(value)
            data.append(row)

        column_widths = [50, 50, 50, 150, 100, 80, 80, 80, 80]

        # Créer le tableau PDF avec les en-têtes ajustés
        table = Table(data, colWidths=column_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        def add_table_with_page_break(doc, table):
            height = sum([element.wrap(doc.width, doc.bottomMargin)[1] for element in elements])
            table_height = table.wrap(doc.width, doc.bottomMargin)[1]

            if height + table_height > doc.pagesize[1] - doc.topMargin - doc.bottomMargin:
                elements.append(PageBreak())  # Nouvelle page

            elements.append(table)

        add_table_with_page_break(doc, table)

        # Construire le document PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    elif format_export == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PVEvaluation.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"

        # Ajouter les en-têtes de titre
        ws.append(['REPOBLIKAN\'I MADAGASIKARA', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['Fitiavana-Tanindrazana-Fandrosoana', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['************', '', '', '', '', '', '', '', '', '', '', '', ''])

        # Ajouter les valeurs de budget
        if data_queryset.exists():
            first_item = data_queryset.first()
            budget_value = first_item.Budget or ''
            exercice_value = first_item.Annee_exercice or ''
            imputation_value = first_item.imputation_administrative or ''
            chapitre_value = first_item.Designation_du_chapitre or ''
            article_value = first_item.libelle_article or ''
            paragraphe_value = first_item.Nom_et_qualite_du_depositaire_comptable or ''

            ws.append(['BUDGET', budget_value])
            ws.append(['EXERCICE', exercice_value])
            ws.append(['IMPUTATION ADMINISTRATIVE', imputation_value])
            ws.append(['CHAPITRE', chapitre_value, 'ARTICLE', article_value, 'PARAGRAPHE', paragraphe_value])
            ws.append([''])  # Ligne vide

        ws.append(field_headers)  # En-têtes complètes

        for item in data_queryset:
            row = [getattr(item, field) for field in fields]
            ws.append(row)

        wb.save(response)
        return response
    else:
        return HttpResponse(f"Format non supporté: '{format_export}'", status=400)


